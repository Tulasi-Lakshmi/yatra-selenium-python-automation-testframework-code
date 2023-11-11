import inspect
import logging
import softest
import csv
from openpyxl import Workbook, load_workbook
class Utils(softest.TestCase):
    def assertListItemText(self,list1,value):
        for stop in list1:
            print("The text is : " + stop.text)
            self.soft_assert(self.assertEqual,stop.text, value)
            if stop.text == value:
                print("Test passed")
            else:
                print("Test Failed")
        self.assert_all()

    def custom_logger(logLevel=logging.DEBUG):
        # set class/method name from where it's called
        logger_name = inspect.stack()[1][3]

        # create logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logging.DEBUG)
        # create console handler or file handler and set the log level
        fh1 = logging.FileHandler("auto11.log", mode='a')
        # create formatter -- how you want your logs to be located
        formatter2 = logging.Formatter('%(asctime)s - %(levelname)s - %(name)s: %(message)s', datefmt='%m/%d/%Y %I:%M:%S %p')
        # add formatter to console or file handler
        fh1.setFormatter(formatter2)
        # add console handler to logger
        logger.addHandler(fh1)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        ws1 = wb[sheet]
        row_count = ws1.max_row
        col_count = ws1.max_column
        # print(row_count)
        # print(col_count)
        for i in range(2, row_count + 1):
            row = []
            for j in range(1, col_count + 1):
                row.append(ws1.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(file_name):
        #Creating an empty ist

        datalist = []

        # Open CSV file
        csvdata = open(file_name,"r")

        # Create csv reader
        reader = csv.reader(csvdata)
        # skip the header
        next(reader)
        # Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist


